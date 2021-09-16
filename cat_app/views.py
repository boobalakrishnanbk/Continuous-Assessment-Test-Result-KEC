from django.db.models.expressions import Exists
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

from cat_app.models import Mark

import openpyxl
from openpyxl import Workbook

# Create your views here.
@login_required(login_url='/login/')
def staff(request):

    if request.method == 'POST':
        if request.POST['file_type'] == "mark" and request.POST:
            mark_file = request.FILES['mark_file']
            # print(Mark.objects.filter(semester=request.POST['semester'],cat=request.POST['cat']).count())
            if Mark.objects.filter(semester=request.POST['semester'],cat=request.POST['cat']).count()==0:
                importMark(mark_file, request)
            else:
                Mark.objects.filter(semester=request.POST['semester'],cat=request.POST['cat']).delete()
                importMark(mark_file, request)
           
    return render(request, 'staff_page.html', {})


# import to database - Mark
def importMark(mark_file, request):
    workbook = openpyxl.load_workbook(mark_file, read_only=True)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)
    students = []
    marks = []
    # if Mark.objects.filter(cat=request.POST['cat'],semester=request.POST['semester']).distinct().count()==0:
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        roll_no = row[0].value
        name = row[1].value
        phone = row[2].value

        for i in range(4, worksheet.max_column,2):
            mark = Mark()
            mark.roll_number = roll_no
            mark.name = name
            mark.phone = phone

            mark.subject_name = row[i].value
            mark.mark = row[i+1].value
            mark.semester = request.POST['semester']
            mark.cat = request.POST['cat']
            marks.append(mark)
        
    Mark.objects.bulk_create(marks)


# homepage
def home(request):
    if request.user.is_authenticated:
        return redirect('/staff/')
    return render(request, 'home.html', {})



def studentLogin(request):
    error = False   
    data = {'semester':[],'cat':[]}
    
    mark =  Mark.objects.all()
    data['semester'] = mark.values('semester').distinct()
    data['cat'] = mark.values('cat').distinct()

    if request.POST['phone'] and request.POST['roll_number']:
        marks = Mark.objects.filter(roll_number=request.POST['roll_number'].upper(),phone=request.POST['phone']).count()
        if marks>1:
            return render(request, 'show.html', {"roll_number":request.POST['roll_number'],"show":"d-none",'data':data})
    else:
        error = True
    return render(request, 'home.html', {'error':True,'data':data})

def fetch_marks(request):

    data = {'semester':[],'cat':[]}
    mark =  Mark.objects.all()
    data['semester'] = mark.values('semester').distinct()
    data['cat'] = mark.values('cat').distinct()
    if request.POST['semester'] and request.POST['cat']:
        marks = Mark.objects.filter(roll_number=request.POST['roll_number'].upper(),semester=request.POST['semester'],cat=request.POST['cat'])
        if marks.count()>0:
            subs_marks = {}
            for i in marks:
                if not i.subject_name == "ATT":
                    if not i.mark == None:
                        subs_marks[i.subject_name] = i.mark
                else:
                    att = i.mark
            try:
                val = marks.values("cat").distinct()[0]['cat']
                cgpa = 0
                gpa= 0
                for i in marks:
                    if not i.subject_name == "GPA":
                        if not i.mark == None:
                            gpa = i.mark
                    elif not i.subject_name == "CGPA":
                        if not i.mark == None:
                            cgpa = i.mark
                    
            except ValueError:
                pass

            return render(request, 'show.html', {
                "roll_number":request.POST['roll_number'],
                "show":"",
                "name":marks.values("name").distinct()[0],
                "semester":marks.values("semester").distinct()[0],
                "exam": marks.values("cat")=="ese" and "End Semester Examination" or ("CAT - "+str(marks.values("cat").distinct()[0]['cat'])),
                "sub_marks":subs_marks,
                "attendance":att,
                "data":data,
                "gpa":gpa,
                "cgpa":cgpa,
            })

        else:
            data = {'semester':[],'cat':[]}
            mark =  Mark.objects.all()
            data['semester'] = mark.values('semester').distinct()
            data['cat'] = mark.values('cat').distinct()
            return render(request, 'show.html', {
                "roll_number":request.POST['roll_number'],
                "show":"d-none",
                'data':data
            })

    
